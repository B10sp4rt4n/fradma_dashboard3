"""
schema_generator.py
Genera plantillas descargables (CSV / XLSX) para cada esquema registrado.

Funciones principales:
    generate_csv_template(schema_id, output_path=None)  -> str con CSV o bytes
    generate_xlsx_template(schema_id, output_path=None) -> bytes (placeholder)
    get_template_columns(schema_id)                     -> list de columnas
    get_sample_rows(schema_id)                          -> list de dicts
"""

import csv
import io
from typing import Optional, Union

from .schema_registry import get_schema
from .sample_data_generator import get_sample_rows as _sample_rows


# =====================================================================
# HELPERS
# =====================================================================

def get_template_columns(schema_id: str) -> list:
    """
    Retorna las columnas sugeridas para la plantilla de un esquema.
    Orden: obligatorios -> recomendados -> opcionales.

    Args:
        schema_id: ID del esquema

    Returns:
        Lista de nombres canonicos de columnas
    """
    schema = get_schema(schema_id)
    cols = (
        schema.get("campos_obligatorios", [])
        + schema.get("campos_recomendados", [])
        + schema.get("campos_opcionales", [])
    )
    # Eliminar duplicados preservando orden
    seen = set()
    result = []
    for c in cols:
        if c not in seen:
            seen.add(c)
            result.append(c)
    return result


def generate_csv_template(
    schema_id: str,
    output_path: Optional[str] = None,
) -> Union[str, bytes]:
    """
    Genera una plantilla CSV para el esquema dado.

    Args:
        schema_id:   ID del esquema (e.g. 'ventas_comercial_v1')
        output_path: Si se especifica, guarda el archivo en esa ruta.
                     Si es None, retorna el contenido como string.

    Returns:
        str con contenido CSV si output_path es None,
        bytes (vacío) si se guardo en disco.
    """
    columns  = get_template_columns(schema_id)
    rows     = _sample_rows(schema_id)

    # Si no hay filas de ejemplo, crear una fila en blanco con encabezados
    if not rows:
        rows = [{col: "" for col in columns}]

    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=columns,
        extrasaction="ignore",
        lineterminator="\n",
    )
    writer.writeheader()
    for row in rows:
        # Solo incluir columnas declaradas; rellenar faltantes con ""
        safe_row = {col: row.get(col, "") for col in columns}
        writer.writerow(safe_row)

    csv_content = output.getvalue()

    if output_path:
        with open(output_path, "w", encoding="utf-8", newline="") as f:
            f.write(csv_content)
        return b""

    return csv_content


def generate_xlsx_template(
    schema_id: str,
    output_path: Optional[str] = None,
) -> bytes:
    """
    Genera una plantilla XLSX para el esquema dado.

    NOTA: Requiere openpyxl. Si no esta disponible, lanza ImportError
    con instruccion de instalacion.

    Args:
        schema_id:   ID del esquema
        output_path: Si se especifica, guarda el archivo en esa ruta.

    Returns:
        bytes del archivo XLSX si output_path es None,
        bytes vacíos si se guardo en disco.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as exc:
        raise ImportError(
            "openpyxl es necesario para generar plantillas XLSX. "
            "Instala con: pip install openpyxl"
        ) from exc

    columns = get_template_columns(schema_id)
    rows    = _sample_rows(schema_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    schema = get_schema(schema_id)
    ws.title = schema.get("nombre", schema_id)[:31]

    # Header row con estilo
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center")

    # Filas de ejemplo
    if not rows:
        rows = [{col: "" for col in columns}]
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    # Ajustar anchos
    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = max(len(col_name) + 4, 14)

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    if output_path:
        with open(output_path, "wb") as f:
            f.write(xlsx_bytes)
        return b""

    return xlsx_bytes
